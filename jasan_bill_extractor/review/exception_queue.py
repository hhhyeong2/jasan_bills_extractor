"""
review/exception_queue.py
---------------------------------
ReviewQueueBuilder (spec.md §2, §6 Phase 4): 저신뢰/검증실패/미매칭 건을 예외 리포트로
모으고, 엑셀 결과 파일에 별도 시트로 하이라이트해 담당자가 예외 건만 확인하면 되게 한다.
"""

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

MIN_CONFIDENCE_DEFAULT = 0.7

HIGHLIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


@dataclass
class ExceptionRow:
    pipeline_row: dict
    reasons: list = field(default_factory=list)


def evaluate(pipeline_row: dict, min_confidence: float = MIN_CONFIDENCE_DEFAULT) -> ExceptionRow:
    reasons = []

    if not pipeline_row.get("site_id"):
        if pipeline_row.get("match_method") not in ("제외(비고지서)", "제외(관할아님/비고지서)"):
            reasons.append("사업장 매칭 실패")

    arith = str(pipeline_row.get("arithmetic_due_status") or "")
    if arith.startswith("불일치"):
        reasons.append(f"산술검증 불일치: {arith}")

    cross = str(pipeline_row.get("cross_field_status") or "")
    if cross == "불일치":
        reasons.append(f"구역간 값 불일치: {pipeline_row.get('cross_field_detail', '')}")

    if pipeline_row.get("history_status") == "이상탐지":
        reasons.append(
            f"전월/전년 대비 이상탐지 (전월대비 {pipeline_row.get('history_prev_month_deviation')}, "
            f"전년동월대비 {pipeline_row.get('history_prev_year_deviation')})"
        )

    conf = pipeline_row.get("min_field_confidence")
    try:
        if conf not in (None, "") and float(conf) < min_confidence:
            reasons.append(f"필드 신뢰도 낮음 ({conf})")
    except (TypeError, ValueError):
        pass

    if pipeline_row.get("error"):
        reasons.append(f"처리 오류: {pipeline_row['error']}")

    return ExceptionRow(pipeline_row=pipeline_row, reasons=reasons)


def build_exception_rows(pipeline_rows: list, min_confidence: float = MIN_CONFIDENCE_DEFAULT) -> list:
    excluded_methods = {"제외(비고지서)", "제외(관할아님/비고지서)"}
    out = []
    for row in pipeline_rows:
        if row.get("match_method") in excluded_methods:
            continue  # 애초에 매칭/검증 대상이 아닌 문서는 예외 큐에도 안 넣음
        ex = evaluate(row, min_confidence)
        if ex.reasons:
            out.append(ex)
    return out


def write_exception_report(exception_rows: list, output_csv: Path) -> None:
    import csv

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "source_file", "frame_index", "doc_index_in_frame", "site_id", "site_name",
        "billing_period", "due_total_amount", "reasons",
    ]
    with open(output_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for ex in exception_rows:
            r = ex.pipeline_row
            writer.writerow({
                "source_file": r.get("source_file"),
                "frame_index": r.get("frame_index"),
                "doc_index_in_frame": r.get("doc_index_in_frame"),
                "site_id": r.get("site_id"),
                "site_name": r.get("site_name"),
                "billing_period": r.get("billing_period"),
                "due_total_amount": r.get("due_total_amount"),
                "reasons": " | ".join(ex.reasons),
            })


def append_exception_sheet(workbook_path: Path, exception_rows: list) -> None:
    """ExcelWriter가 만든 워크북에 '검토필요' 시트를 추가하고 행을 하이라이트한다."""
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb.create_sheet("검토필요")
    header = [
        "source_file", "frame_index", "doc_index_in_frame", "site_id", "site_name",
        "billing_period", "due_total_amount", "사유",
    ]
    ws.append(header)
    for ex in exception_rows:
        r = ex.pipeline_row
        ws.append([
            r.get("source_file"), r.get("frame_index"), r.get("doc_index_in_frame"),
            r.get("site_id"), r.get("site_name"), r.get("billing_period"),
            r.get("due_total_amount"), " | ".join(ex.reasons),
        ])
        for cell in ws[ws.max_row]:
            cell.fill = HIGHLIGHT_FILL
    wb.save(workbook_path)
