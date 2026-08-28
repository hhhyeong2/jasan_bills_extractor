"""
writer/excel_writer.py
---------------------------------
ExcelWriter (spec.md §3, §6 Phase 4): 검증을 통과한 고신뢰 문서를 경기북부 엑셀양식
컬럼에 맞춰 새 워크북에 기입한다.

입력은 두 가지다:
  1) 경기북부 엑셀양식 파일(기존 이력 포함) - 헤더 구조 + 사업장별 참고 데이터(사업장명,
     직전 사용기간) 소스로 사용
  2) 고지서에서 추출/검증된 문서 데이터 (scripts.run_pipeline 결과)

중요: 사용자의 원본 템플릿은 절대 직접 덮어쓰지 않는다. 템플릿을 읽어 헤더만 복사한 새
워크북을 만들고, 거기에 데이터 행을 추가한다.

컬럼 매핑 근거:
    D  사업장명            <- 엑셀양식에 이미 기재된 해당 사업장번호의 사업장명을 그대로 사용
                              (엑셀 표기 컨벤션 유지를 위해 SiteMatcher 결과보다 우선함).
                              엑셀에 해당 사업장번호 이력이 전혀 없는 신규 사업장만
                              SiteMatcher가 넘겨준 이름으로 대체.
    E  사업장번호(필수)     <- SiteMatcher 결과
    F  사용기간(START)      <- 고지서에 사용기간이 명확히 있으면 그 값. 없으면 엑셀양식에
    G  사용기간(END)           있는 해당 사업장의 가장 최근 사용기간에서 '월'만 +1 (12월이면
                              다음 해 1월로 롤오버, '일'은 그대로 유지)
    H  청구년월(필수)       <- 엑셀양식에 있는 해당 사업장의 최근 청구년월에서 월만 +1
                              (F/G 사용기간과 동일한 방식. 이력이 없는 신규 사업장만
                              고지서에서 추출한 billing_period로 대체)
    I  납기년월(필수)       <- 엑셀양식에 있는 해당 사업장의 최근 납기년월에서 월만 +1
                              (이력이 없으면 위에서 계산한 H값과 동일)
    J  지급일자(필수)       <- 오늘 날짜 기준 +3영업일(토/일 제외, 공휴일은 미반영)
    Q  청구요금계           <- due_total_amount
    T  전기요금             <- supply_amount (없으면 current_period_charge로 대체:
                              관리비고지서는 세부항목이 분해되지 않고 한 줄 합계로만 나오는
                              경우가 많음 - poc_out/poc_results_review.csv 검수 결과 참고)
    U  부가가치세           <- vat_amount
    W  전력기금             <- FieldCombiner 합산값 (전력기금원본+연체료+미납액+미납연체료)
    X  기타요금(청소용역비)  <- other_fee

A(지급월)/C(유형)/Y(납기일)/Z~AF(계좌·세금계산서 정보)는 고지서 이미지만으로 신뢰성 있게
판단할 수 없는 회계 프로세스/계좌 정보라 자동으로 채우지 않는다(계좌 정보를 잘못 자동
기입하면 자금이 잘못된 곳으로 나갈 위험이 있어 특히 더 위험함). A/C는 최빈값(지급월='월',
유형='A' 또는 세금계산서類는 'B')으로 기본값만 채워두고, 나머지는 빈 칸으로 남겨 담당자가
채우도록 한다.
"""

import calendar
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional, Union

import openpyxl

from jasan_bill_extractor.mapping.field_combiner import combine_power_fund

COLUMN_ORDER = [
    "지급월", "계좌명", "유형", "사업장명", "사업장번호(필수)",
    "사용기간(START)", "사용기간(END)", "청구년월(필수)", "납기년월(필수)", "지급일자(필수)",
    "지침1", "지침2", "지침3", "지침4", "지침5", "전력사용량(kWh)",
    "청구요금계", "기본요금", "사용량요금", "전기요금", "부가가치세", "원단위절사",
    "전력기금", "기타요금(청소용역비)", "납기일(필수)",
    "은행코드", "계좌번호", "예금주명", "세금계산서작성일", "전자증빙구분", "전자세금계산서승인번호",
    "비고(단가)",
]

_TAX_INVOICE_TYPES = {"세금계산서", "계산서"}

# 경기북부 엑셀양식 열 위치 (0-indexed): D=3 사업장명, E=4 사업장번호, F=5 사용기간START,
# G=6 사용기간END, H=7 청구년월, I=8 납기년월
_COL_SITE_NAME, _COL_SITE_ID, _COL_USAGE_START, _COL_USAGE_END, _COL_BILLING_PERIOD, _COL_DUE_PERIOD = 3, 4, 5, 6, 7, 8


def _default_type_code(doc_type: Optional[str]) -> str:
    return "B" if doc_type in _TAX_INVOICE_TYPES else "A"


def _parse_date(value: Union[str, datetime, date, None]) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def add_one_month(d: date) -> date:
    """월만 +1 (12월이면 다음 해 1월로 롤오버), 일(day)은 그대로 유지하되 대상 월의
    말일을 넘지 않도록 clamp."""
    if d.month == 12:
        year, month = d.year + 1, 1
    else:
        year, month = d.year, d.month + 1
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, min(d.day, last_day))


def add_one_month_yyyymm(period: str) -> str:
    """'YYYYMM' 문자열의 월만 +1 (12월이면 다음 해 1월로 롤오버)."""
    year, month = int(period[:4]), int(period[4:6])
    if month == 12:
        year, month = year + 1, 1
    else:
        month += 1
    return f"{year:04d}{month:02d}"


def add_business_days(start: date, n: int) -> date:
    """토/일요일만 건너뛴 영업일 계산 (공휴일 캘린더는 반영하지 않음)."""
    d = start
    added = 0
    while added < n:
        d = d + timedelta(days=1)
        if d.weekday() < 5:  # 0=월 ... 4=금
            added += 1
    return d


@dataclass
class SiteReference:
    site_name: str
    usage_start: Optional[date]
    usage_end: Optional[date]
    billing_period: str
    due_period: str


def load_reference_data(template_path: Path) -> dict:
    """경기북부 엑셀양식에서 사업장번호별로 청구년월이 가장 최신인 행을 골라
    {사업장번호: SiteReference}로 반환한다. (D 사업장명 표기, F/G 사용기간 롤오버 기준값 소스)"""
    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb.active
    ref: dict[str, SiteReference] = {}
    latest_period: dict[str, int] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= _COL_BILLING_PERIOD:
            continue
        site_name = row[_COL_SITE_NAME]
        site_id = row[_COL_SITE_ID]
        if not site_id or not site_name:
            continue
        site_id = str(site_id).strip()
        billing_period = row[_COL_BILLING_PERIOD]
        try:
            bp_int = int(billing_period) if billing_period else 0
        except (ValueError, TypeError):
            bp_int = 0

        if site_id in latest_period and bp_int < latest_period[site_id]:
            continue  # 이미 더 최신 행을 찾았으면 건너뜀
        latest_period[site_id] = bp_int
        due_period = row[_COL_DUE_PERIOD] if len(row) > _COL_DUE_PERIOD else None
        ref[site_id] = SiteReference(
            site_name=str(site_name).strip(),
            usage_start=_parse_date(row[_COL_USAGE_START]),
            usage_end=_parse_date(row[_COL_USAGE_END]),
            billing_period=str(billing_period) if billing_period else "",
            due_period=str(due_period) if due_period else "",
        )
    return ref


def build_row(
    doc: dict,
    site_id: str,
    fallback_site_name: str,
    reference: Optional[SiteReference],
    payment_date: date,
) -> dict:
    combined = combine_power_fund(doc)
    supply_or_aggregate = doc.get("supply_amount")
    if supply_or_aggregate is None:
        supply_or_aggregate = doc.get("current_period_charge")

    # D: 사업장명 - 엑셀양식에 이력이 있으면 그 표기를 그대로 사용, 없으면(신규 사업장)
    # SiteMatcher가 넘겨준 이름으로 대체
    site_name = reference.site_name if reference else fallback_site_name

    # F/G: 사용기간 - 고지서에 시작/종료 둘 다 명확히 있어야만 그 값 쌍으로 덮어쓴다.
    # 한쪽만 있으면 "명확하게 있다"고 볼 수 없으므로(시작만/종료만 있으면 두 값이 서로
    # 안 맞는 조합이 될 위험) 둘 다 기존 양식의 최근 사용기간에서 월만 +1 한 값으로 통일한다.
    doc_usage_start = _parse_date(doc.get("usage_start_date"))
    doc_usage_end = _parse_date(doc.get("usage_end_date"))
    if doc_usage_start is not None and doc_usage_end is not None:
        usage_start, usage_end = doc_usage_start, doc_usage_end
    else:
        usage_start = add_one_month(reference.usage_start) if reference and reference.usage_start else None
        usage_end = add_one_month(reference.usage_end) if reference and reference.usage_end else None

    # H/I: 청구년월/납기년월 - 고지서에서 읽은 값이 아니라 기존 엑셀양식의 해당 사업장
    # 최근 값에서 월만 +1 (F/G 사용기간과 같은 방식). 이력이 전혀 없는 신규 사업장만
    # 고지서에서 추출한 billing_period로 대체한다.
    doc_billing_period = doc.get("billing_period") or ""
    if reference and reference.billing_period:
        billing_period_calc = add_one_month_yyyymm(reference.billing_period)
    else:
        billing_period_calc = doc_billing_period
    if reference and reference.due_period:
        due_period_calc = add_one_month_yyyymm(reference.due_period)
    else:
        due_period_calc = billing_period_calc

    return {
        "지급월": "월",  # 기본값(최빈값) - 반기/모자분리 등 예외는 담당자 확인 필요
        "계좌명": "",  # 계좌 정보는 자동 기입하지 않음(오지급 위험)
        "유형": _default_type_code(doc.get("doc_type")),
        "사업장명": site_name,
        "사업장번호(필수)": site_id,
        "사용기간(START)": usage_start.isoformat() if usage_start else "",
        "사용기간(END)": usage_end.isoformat() if usage_end else "",
        "청구년월(필수)": billing_period_calc,
        "납기년월(필수)": due_period_calc,
        "지급일자(필수)": payment_date.isoformat(),
        "지침1": "", "지침2": "", "지침3": "", "지침4": "", "지침5": "",
        "전력사용량(kWh)": "",
        "청구요금계": doc.get("due_total_amount") or "",
        "기본요금": "",
        "사용량요금": "",
        "전기요금": supply_or_aggregate if supply_or_aggregate is not None else "",
        "부가가치세": doc.get("vat_amount") or "",
        "원단위절사": "",
        "전력기금": combined.combined_value,
        "기타요금(청소용역비)": doc.get("other_fee") or "",
        "납기일(필수)": "",  # 고지서에서 날짜는 읽히지만 '일(day)'만 구조화 추출 안 됨 - Phase 5 개선 후보
        "은행코드": "", "계좌번호": "", "예금주명": "",
        "세금계산서작성일": "", "전자증빙구분": "", "전자세금계산서승인번호": "",
        "비고(단가)": "",
    }


def write_workbook(template_path: Path, rows: list, output_path: Path) -> int:
    """template_path의 헤더만 복사한 새 워크북에 rows를 기입해 output_path로 저장한다.
    template_path 원본은 절대 수정하지 않는다. 반환값: 기입된 행 수.
    """
    src_wb = openpyxl.load_workbook(template_path)
    src_ws = src_wb.active
    header_row = [cell.value for cell in src_ws[1]]

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = src_ws.title
    out_ws.append(header_row)

    for row in rows:
        out_ws.append([row.get(col, "") for col in COLUMN_ORDER])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_path)
    return len(rows)
